import requests
from bs4 import BeautifulSoup
import openpyxl
import os
from utils import remove_char, create_folder

def create_data(subject_names, subject_ids):
    field_list_codes_and_des = []
    field_codes = []
    field_codes_dict = {}
    for i in range(len(subject_names)):
        chosen_id = subject_ids[i]
        doc_url = "https://sylabus.sggw.edu.pl/pl/document/" + chosen_id + ".jsonHtml"

        payload = {}
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"
        }

        response = requests.request("GET", doc_url, headers=headers, data=payload)
        html = response.json()["html"]
        bs4 = BeautifulSoup(html, "html.parser")

        codes_and_descriptions1 = []
        codes_and_descriptions2 = []
        subject_codes = []

        for item2 in bs4.find_all("span", class_="popup"):
            if item2.get("data-bs-original-title") is not None:
                code = remove_char(item2.get_text().strip(), ",")
                code_description = item2.get("data-bs-original-title")
                codes_and_descriptions1.append((code, code_description))
            if item2.get("title") is not None:
                code = remove_char(item2.get_text().strip(), ",")
                code_description = item2.get("title")
                subject_codes.append(code)
                codes_and_descriptions2.append((code, code_description))
                field_codes.append(code)
        
        subject_codes = list(filter(None, subject_codes))
        field_codes_dict[subject_names[i]] = subject_codes
        #field_codes.append(list(set(subject_codes)))
    
    field_list_codes_and_des = list(set(codes_and_descriptions2))
    field_codes = list(filter(None, list(set(field_codes))))
        
    return field_codes_dict, field_codes


def save_data(subject_names, subject_ids, selected_field):
    
    field_codes_dict, field_codes = create_data(subject_names, subject_ids)
    field_codes = list(filter(lambda x: len(x) != 2, field_codes))
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Wpisz nazwy przedmiotów w pierwszej kolumnie
    sheet["A1"] = "Przedmioty"
    for index, subject in enumerate(subject_names, start=2):
        sheet[f"A{index}"] = subject

    # Utwórz nagłówki kolumn na podstawie unikalnych kodów
    unique_codes = set(code for code in field_codes)
    unique_codes = sorted(unique_codes)
    for index, code in enumerate(unique_codes, start=2):
        sheet[f"{openpyxl.utils.get_column_letter(index)}1"] = code

    # Wypełnij arkusz wartościami
    for col, subject in enumerate(subject_names, start=2):
        codes = field_codes_dict[subject]
        for row, code in enumerate(unique_codes, start=2):
            count = codes.count(code)
            sheet[f"{openpyxl.utils.get_column_letter(row)}{col}"] = count

    
    folder_faculties_name = "Selected_fields_of_study"
    file_path = os.path.join(os.getcwd(), folder_faculties_name, f"{selected_field}.xlsx")
    wb.save(file_path)
