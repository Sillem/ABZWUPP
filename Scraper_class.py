import requests
from bs4 import BeautifulSoup
import streamlit as st
import os
import openpyxl
from time import sleep


def remove_char(input_string, char_to_remove):
    result = ""
    for char in input_string:
        if char != char_to_remove:
            result += char
    return result


class Scraper(object):
    """
    Klasa odpowiedzialna za pobieranie danych z witryny https://sylabus.sggw.edu.pl/

  
    """
    def __init__(self):
        """Konstruktor klasy inicjalizujacy puste listy 'subject_names', 'subject_ids',
         'field_list_codes_and_des', 'field_codes', 'field_codes_dict' ktore beda
        przechowywac nazwy, identyfikatory,kody i opisy przedmiotow do pozniejszego wykorzystania.
        """
        self.subject_names = []
        self.subject_ids = []
        self.field_list_codes_and_des = []
        self.field_codes = []
        self.field_codes_dict = {}

    def reset_codes_data(self):
        """
        Ta funkcja jest funkcja pomocnicza, sluzaca do resetowania pol 'field_list_codes_and_des',
        'field_codes', 'field_codes_dict'. 
        """
        self.field_list_codes_and_des = []
        self.field_codes = []
        self.field_codes_dict = {}

    def create_folder(self, folder_name, path=None):
        """
        Ta funkcja sluzy do tworzenaia folderu, w ktorym beda zapisywane pobrane teksty. 

        Args:
            folder_name (str): string z nazwa folderu
            path (str, optional): string z docelowa sciezka folderu, domyslnie None
        """
        try:
            if path is None:
                path = os.getcwd()  # Domyślna ścieżka to aktualny katalog roboczy

            new_folder_path = os.path.join(path, folder_name)

            if not os.path.exists(new_folder_path):
                os.makedirs(new_folder_path)
                print(
                    f"Folder '{folder_name}' created successfully at: {new_folder_path}"
                )
            else:
                print(f"Folder '{folder_name}' already exists at: {new_folder_path}")
        except Exception as e:
            print(f"An error occurred: {e}")

    def create_data_single(self, html, subject_name):
        """
        Ta funkcja przygotowuje dane ze strony podanej w parametrze funkcji i zapisuje odpowiednio w polach:
        self.field_codes_dict oraz self.field_codes

        Args:
            html (str): string z linkiem do wybranej podstrony
            subject_name (str): string z nazwa wybranego przedmiotu
        """
        bs4 = BeautifulSoup(html, "html.parser")

        codes_and_descriptions1 = []
        codes_and_descriptions2 = []
        subject_codes = []

        for item2 in bs4.find_all("span", class_="popup"):
            if item2.get("data-bs-original-title") is not None:
                code = remove_char(item2.get_text().strip(), ",")
                code_description = item2.get("data-bs-original-title")
                codes_and_descriptions1.append((code, code_description))
            if item2.get("title") is not None:
                code = remove_char(item2.get_text().strip(), ",")
                code_description = item2.get("title")
                subject_codes.append(code)
                codes_and_descriptions2.append((code, code_description))
                self.field_codes.append(code)

        subject_codes = list(filter(None, subject_codes))
        self.field_codes_dict[subject_name] = subject_codes

    def get_effects_content_codes(self, chosen_id, subject_name):
        """
        Ta funkcja pobiera efekty kształcenia i treści programowe dla wybranego przedmiotu. 

        Args:
            chosen_id (int): int z kolejnym nr przedmiotu na danym kierunku
            subject_name (str): string z wybrana nazwa przedmiotu

        Returns:
            tuple codes_and_descriptions, learning_effects, course_content: krotka ze slownikami z kodami,
            efektami uczenia sie i tresciami programowymi
        """
        payload = {}
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"
        }
        doc_url = "https://sylabus.sggw.edu.pl/pl/document/" + chosen_id + ".jsonHtml"
        for i, trail in enumerate(range(3)):
            try:
                response = requests.request(
                    "GET", doc_url, headers=headers, data=payload
                )
                break
            except (
                requests.exceptions.ConnectTimeout,
                requests.exceptions.ConnectionError,
            ) as e:
                print("Timeout. Retrying...")
                sleep(0.1)
        html = response.json()["html"]
        self.create_data_single(html, subject_name)

        bs4 = BeautifulSoup(html, "html.parser")
        course_content = ""
        learning_effects = ""

        h1s = bs4.find_all("h1")
        for h1 in h1s:
            if h1.get_text().strip() == "Efekty uczenia się dla przedmiotu":
                efekty_tag = h1.find_next()

                for i, item in enumerate(efekty_tag.find_all("td", class_="code")):
                    if (
                        "row_code" not in item["class"]
                        and "row_header" not in item["class"]
                    ):
                        if item.find("span") is None:
                            learning_effects += (
                                item.find_next().get_text().strip() + " "
                            )

            if h1.get_text().strip() == "Treści programowe":
                tresc_tag = h1.find_next()

                for i, item in enumerate(tresc_tag.find_all("td", class_="code")):
                    if (
                        "row_code" not in item["class"]
                        and "row_header" not in item["class"]
                    ):
                        the_text = item.find_next().get_text().strip()
                        course_content += the_text + " "
        bs = BeautifulSoup(html, "html.parser")
        codes_and_descriptions = {}
        for item in bs.find_all("span", class_="popup"):
            if item.get("title") is not None:
                code = remove_char(item.get_text().strip(), ",")
                code_description = item.get("title")
                codes_and_descriptions[code] = code_description
            else:
                print("No popup title found")

        return codes_and_descriptions, learning_effects, course_content

    def save_data(self, selected_field):
        """
        Ta funkcja zapisuje pobrane dane na temat wybranego kierunku do pliku .xlsx.

        Args:
            selected_field (str): string z nazwa wybranego kierunku
        """
        field_codes_dict = self.field_codes_dict
        field_codes = list(filter(None, list(set(self.field_codes))))

        field_codes = list(filter(lambda x: len(x) != 2, field_codes))
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Wpisywanie nazwy przedmiotów w pierwszej kolumnie
        sheet["A1"] = "Przedmioty"
        for index, subject in enumerate(self.subject_names, start=2):
            sheet[f"A{index}"] = subject

        # Tworzenie nagłówków z kodów studiów
        unique_codes = set(code for code in field_codes)
        unique_codes = sorted(unique_codes)
        for index, code in enumerate(unique_codes, start=2):
            sheet[f"{openpyxl.utils.get_column_letter(index)}1"] = code

        # Wypełnianie arkusza kalkulacyjnego wartościami
        for col, subject in enumerate(self.subject_names, start=2):
            codes = field_codes_dict[subject]
            for row, code in enumerate(unique_codes, start=2):
                count = codes.count(code)
                sheet[f"{openpyxl.utils.get_column_letter(row)}{col}"] = count
        current_path = os.path.dirname(__file__)
        default_path = os.path.abspath(os.path.join(current_path, os.pardir))
        field_names_folder = "Selected_fields_of_study"
        folder_path = os.path.join(default_path, field_names_folder)
        print("co to ścieżka: " +str(type(folder_path)))
        self.create_folder(f"{selected_field}", folder_path)
        file_path = os.path.join(folder_path, selected_field, f"{selected_field}.xlsx")
        wb.save(file_path)

    def get_data(self, sub_sub_url, progress_bar):
        """
        Ta funkcja tworzy slowniki z kodami, efektami nauczania i tresciami programowymi dla wybranego przez
        uzytkownika kierunku nauczania. 

        Args:
            sub_sub_url (str): string z linkiem do podstrony wybranego kierunku nauczania
            progress_bar (class 'streamlit.delta_generator.DeltaGenerator): klasa z modulu streamlit do 
            tworzenia paska postepu

        Returns:
            tuple effects, contents, codes: krotka ze slownikami efektow, tresci i kodow dla wybranego 
            kierunku
        """
        self.reset_codes_data()
        # Pobieranie listy przedmiotów z wybranego kierunku (ze strony z sylabusami)
        the_class = "syl-get-document syl-pointer"
        payload = {}
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"
        }
        for i, trail in enumerate(range(3)):
            try:
                response = requests.request(
                    "GET", sub_sub_url, headers=headers, data=payload
                )
                print("Sending request #", i, " to: ", sub_sub_url)
                break
            except (
                requests.exceptions.ConnectTimeout,
                requests.exceptions.ConnectionError,
            ) as e:
                print("Timeout. Retrying...")
                sleep(1)
        bs3 = BeautifulSoup(response.content, "html.parser")

        subject_divs = bs3.find_all(class_=the_class)
        subject_names = []
        subject_ids = []
        n = 0

        for i, item in enumerate(subject_divs):
            subject_names.append(item.string.strip())
            subject_ids.append(item.get("id"))
        self.subject_names = subject_names
        self.subject_ids = subject_ids
        # stare get_dictionaries
        effects = {}
        contents = {}
        codes = {}
        for index, i in enumerate(self.subject_ids):
            returned_values = self.get_effects_content_codes(i, subject_names[index])
            codes_and_descriptions = returned_values[0]
            learning_effects = returned_values[1]
            course_content = returned_values[2]

            effects[self.subject_names[index]] = learning_effects
            contents[self.subject_names[index]] = course_content
            codes[self.subject_names[index]] = codes_and_descriptions
            if index % 2:
                progress_bar.progress(n + 1)
                n += 1
        return effects, contents, codes

    def get_description(self, selected_field, sub_sub_url):
        """
        Ta funkcja uzyskuje opis kierunku nauczania wybranego przez uzytkownika. 

        Args:
            selected_field (str): string z nazwa wybranego kierunku studiow
            sub_sub_url (str): string z linkiem do podstrony wybranego kierunku studiow
        """
        st.write(f"Wybrany kierunek: {selected_field}")
        the_class = "syl-grid-tab-content tab-pane fade active show"
        payload = {}
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"
        }
        for i, trail in enumerate(range(3)):
            try:
                response = requests.request(
                    "GET", sub_sub_url, headers=headers, data=payload
                )
                print("Sending request #", i, " to: ", sub_sub_url)
                break
            except (
                requests.exceptions.ConnectTimeout,
                requests.exceptions.ConnectionError,
            ) as e:
                print("Timeout. Retrying...")
                sleep(0.1)
        bs3 = BeautifulSoup(response.content, "html.parser")
        paragraphs = bs3.find("div", {"id": "syl-grid-period-info"}).find_all("p")

        for paragraph in paragraphs:
            paragraph_text = paragraph.get_text(strip=True).replace(">", "")
            if paragraph_text != "" and paragraph_text != "Program studiów":
                st.write(paragraph_text)
